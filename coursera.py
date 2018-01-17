#!/usr/bin/env python3


import random
import requests
import bs4
import openpyxl
import os
import argparse
from lxml import objectify


def fetch_response_of_get_request_as_bytecode(url):
    response = requests.get(url)
    response_data = response.content
    return response_data


def parse_xml_with_full_courses_list(coursera_courses_xml_as_bytecode):
    urlset_xml_root_object = objectify.fromstring(
        coursera_courses_xml_as_bytecode
    )
    full_courses_list = [
        url_object.loc.text
        for url_object in urlset_xml_root_object.iterchildren()
    ]
    return full_courses_list


def choose_random_courses(amount_of_courses, full_courses_list):
    courses_links_list = random.sample(
        full_courses_list,
        amount_of_courses
    )
    return courses_links_list


def grab_bs4_objects_from_html_page(course_page_html):
    grabbed_page = bs4.BeautifulSoup(course_page_html, "html.parser")
    language_raw_object = grabbed_page.find(
        "div",
        class_="rc-Language"
    )
    start_date_raw_object = grabbed_page.find(
        "div",
        class_="startdate rc-StartDateString caption-text"
    )
    course_program_raw_object = grabbed_page.find(
        "div",
        class_="rc-WeekView"
    )
    rating_raw_object = grabbed_page.find(
        "div",
        class_="ratings-text bt3-hidden-xs"
    )
    course_name_raw_object = grabbed_page.find(
        "h1",
        class_="title display-3-text"
    )
    course_ratings_raw_object = grabbed_page.find(
        "div",
        class_="ratings-text headline-2-text"
    )
    raw_page_data = [
        course_name_raw_object,
        language_raw_object,
        start_date_raw_object,
        course_program_raw_object,
        course_ratings_raw_object
    ]
    return raw_page_data


def get_raw_courses_data(courses_links_list):
    raw_courses_data = [(
        "url", "name",
        "language",
        "date",
        "duration(weeks)",
        "rating"
    )]
    for course_link in courses_links_list:
        course_page_html = fetch_response_of_get_request_as_bytecode(
            course_link
        )
        raw_page_data = grab_bs4_objects_from_html_page(
            course_page_html
        )
        raw_page_data_with_url = [course_link]
        raw_page_data_with_url.extend(raw_page_data)
        raw_courses_data.append(raw_page_data_with_url)
    return raw_courses_data


def convert_raw_courses_data_to_excel_workbook(raw_courses_data):
    excel_workbook = openpyxl.Workbook()
    work_sheet = excel_workbook.active
    work_sheet.title = "some_coursera_offers"
    column_offset = 1
    row_offset = 1
    header_fill = openpyxl.styles.PatternFill(
        patternType="solid",
        fgColor="0000FF00"
    )
    header_font = openpyxl.styles.Font(
        name="FreeMono",
        size=13,
        bold=True,
        italic=False,
        vertAlign=None,
        underline=None,
        strike=False,
        color='FF000000'
    )
    regular_font = openpyxl.styles.Font(
        name="FreeMono",
        size=10,
        bold=False,
        italic=False,
        vertAlign=None,
        underline=None,
        strike=False,
        color='FF000000'
    )
    for row_id, row in enumerate(raw_courses_data, start=row_offset):
        for column_id, cell_input_data in enumerate(row, start=column_offset):
            cell = work_sheet.cell(column=column_id, row=row_id)
            if isinstance(cell_input_data, str):
                cell.value = cell_input_data
            elif cell_input_data is None:
                cell.value = "N/A"
            elif (isinstance(cell_input_data, bs4.element.Tag) and
                    cell_input_data.attrs["class"] == ["rc-WeekView"]):
                cell.value = len(cell_input_data)
            else:
                cell.value = cell_input_data.get_text()
    return excel_workbook


def create_cli_parser_and_collect_cli_arguments():
    default_courses_amount = 10
    max_courses_amount = 100
    cli_parser = argparse.ArgumentParser(
        description=(
            "Программа для формирования списка курсов с www.coursera.org"
        )
    )
    cli_parser.add_argument(
        "--directory",
        "-d",
        type=str,
        dest="path_to_directory",
        metavar="target directory path",
        default="./",
        help=(
            "путь до директории, "
            "куда нужно будет сохранить файл с результатами"
        )
    )
    cli_parser.add_argument(
        "--filename",
        "-f",
        type=str,
        dest="file_name",
        metavar="target file name",
        default="coursera.xlsx",
        help="имя файла, куда нужно будет сохранить результаты"
    )
    cli_parser.add_argument(
        "--amount",
        "-a",
        type=int,
        dest="amount_of_courses",
        metavar="amount of courses",
        default=default_courses_amount,
        choices=range(1, max_courses_amount),
        help=(
            "количество курсов, "
            "которые нужно будет рассмотреть и записать"
        )
    )
    cli_arguments = cli_parser.parse_args()
    return cli_arguments


if __name__ == '__main__':
    cli_arguments = create_cli_parser_and_collect_cli_arguments()
    path_to_directory = cli_arguments.path_to_directory
    file_name = cli_arguments.file_name
    amount_of_courses = cli_arguments.amount_of_courses
    if not os.path.exists(path_to_directory):
        sys.exit(
            "Такая директория не существует"
            "\n"
            "Перезапустите код, указав корректную директорию "
            "или не указывайте совсем"
        )
    xml_with_full_courses_list = fetch_response_of_get_request_as_bytecode(
        "https://www.coursera.org/sitemap~www~courses.xml"
    )
    full_courses_list = parse_xml_with_full_courses_list(
        xml_with_full_courses_list
    )
    courses_links_list = choose_random_courses(
        amount_of_courses,
        full_courses_list
    )
    raw_courses_data = get_raw_courses_data(courses_links_list)
    excel_workbook_file_path = os.path.join(
        path_to_directory,
        file_name
    )
    if excel_workbook_file_path.endswith("xlsx"):
        excel_workbook_file_path_with_extension = excel_workbook_file_path
    else:
        excel_workbook_file_path_with_extension = "{}.xlsx".format(
            excel_workbook_file_path
        )
    excel_workbook = convert_raw_courses_data_to_excel_workbook(
        raw_courses_data
    )
    excel_workbook.save(excel_workbook_file_path_with_extension)
    print(
        "\n"
        "Данные о тренингах Coursera записаны в файл {}".format(
            excel_workbook_file_path_with_extension
        )
    )
